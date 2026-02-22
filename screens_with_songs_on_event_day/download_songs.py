#!/usr/bin/env python3
"""
Download audio from YouTube for Music Bingo.
Reads sources/music_bingo_songs.xlsx: column 1 = song name, column 2 = interval(s), column 3 = YouTube URL.
Extracts the given time intervals from each video and saves to output/<FOLDER_NAME>/<song_name>.m4a.
"""

import argparse
import re
import subprocess
import sys
import tempfile
import time
from pathlib import Path

import openpyxl
import yt_dlp

SCRIPT_DIR = Path(__file__).resolve().parent
SOURCES_XLSX = SCRIPT_DIR / "sources" / "music_bingo_songs.xlsx"
OUTPUT_BASE = SCRIPT_DIR / "output"

# Characters unsafe in filenames (Windows + common FS)
UNSAFE_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

PROGRESS_BAR_WIDTH = 32
DOWNLOAD_RETRIES = 3
DOWNLOAD_RETRY_DELAY_SEC = 8
DOWNLOAD_PAUSE_BETWEEN_SEC = 2


def _format_size(n: float) -> str:
    if n < 1024:
        return f"{n:.0f} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} KiB"
    return f"{n / (1024 * 1024):.1f} MiB"


def _progress_hook(d: dict, stream: bool = True) -> None:
    """yt-dlp progress hook: print a one-line progress bar."""
    if d.get("status") != "downloading":
        if d.get("status") == "finished" and stream:
            sys.stdout.write("\r" + " " * 60 + "\r")
            sys.stdout.flush()
        return
    down = d.get("downloaded_bytes") or 0
    total = d.get("total_bytes")
    if total and total > 0:
        pct = min(100, 100 * down / total)
        filled = int(PROGRESS_BAR_WIDTH * pct / 100)
        bar = "[" + "#" * filled + "-" * (PROGRESS_BAR_WIDTH - filled) + "]"
        sys.stdout.write(f"\r  {bar} {pct:5.1f}%  {_format_size(down)} / {_format_size(total)}   ")
    else:
        bar = "[" + "?" * PROGRESS_BAR_WIDTH + "]"
        sys.stdout.write(f"\r  {bar}  ---%  {_format_size(down)}   ")
    sys.stdout.flush()


def sanitize_filename(name: str) -> str:
    """Replace unsafe characters so the string is safe as a filename."""
    return UNSAFE_FILENAME.sub("_", name).strip() or "unnamed"


def parse_interval(s: str) -> list[tuple[float, float]]:
    """
    Parse interval string into (start_sec, end_sec) pairs.
    Accepts: "0:30-1:00", "30-60", "0:30-1:00, 1:30-2:00".
    Returns list of (start, end) in seconds.
    """
    def to_seconds(x: str) -> float:
        x = x.strip()
        if ":" in x:
            parts = x.split(":")
            if len(parts) == 2:
                return int(parts[0], 10) * 60 + float(parts[1].replace(",", "."))
            if len(parts) == 3:
                return int(parts[0], 10) * 3600 + int(parts[1], 10) * 60 + float(parts[2].replace(",", "."))
        return float(x.replace(",", "."))

    if not s or not str(s).strip():
        return []
    out = []
    for part in str(s).split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            start = to_seconds(a)
            end = to_seconds(b)
            if end > start:
                out.append((start, end))
    return out


def load_rows(xlsx_path: Path, sheet_index: int = 0) -> list[tuple[str, str, str]]:
    """Load (col1_song_name, col2_intervals, col3_youtube_url). Skip header row. Skip rows with empty URL."""
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Source file not found: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
        wb.close()
        raise ValueError(f"Sheet index {sheet_index} out of range. Sheets: {list(wb.sheetnames)}")
    ws = wb.worksheets[sheet_index]
    rows = list(ws.iter_rows(min_col=1, max_col=3, values_only=True))
    wb.close()
    result = []
    for i, row in enumerate(rows):
        if i == 0:
            continue  # header
        c1, c2, c3 = (row[0], row[1] if len(row) > 1 else None, row[2] if len(row) > 2 else None)
        name = (c1 and str(c1).strip()) or ""
        intervals = (c2 and str(c2).strip()) or ""
        url = (c3 and str(c3).strip()) or ""
        if not url or not name:
            continue
        if "youtube.com" in url or "youtu.be" in url:
            result.append((name, intervals, url))
    return result


def download_audio(url: str, out_dir: Path, ffmpeg_location: str | None = None, progress: bool = True) -> Path | None:
    """Download audio from YouTube into out_dir. Returns path to the downloaded file (e.g. .m4a) or None."""
    opts = {
        "format": "bestaudio[ext=m4a]/bestaudio",
        "outtmpl": str(out_dir / "audio.%(ext)s"),
        "quiet": True,
        "no_warnings": True,
        "postprocessors": [{"key": "FFmpegExtractAudio", "preferredcodec": "m4a", "preferredquality": "0"}],
    }
    if progress:
        opts["progress_hooks"] = [_progress_hook]
    if ffmpeg_location:
        opts["ffmpeg_location"] = ffmpeg_location
    try:
        with yt_dlp.YoutubeDL(opts) as ydl:
            ydl.download([url])
        # Postprocessor may rename to .m4a
        for f in out_dir.iterdir():
            if f.suffix.lower() in (".m4a", ".mp3", ".webm", ".opus"):
                return f
        return None
    except Exception as e:
        err = str(e)
        if "ffprobe" in err.lower() or "ffmpeg" in err.lower():
            print("  ffmpeg/ffprobe not found. Install: sudo apt install ffmpeg  (or use --ffmpeg-location)", file=sys.stderr)
        print(f"  yt-dlp error: {e}", file=sys.stderr)
        return None


def _ffmpeg_cmd(ffmpeg_location: str | None) -> str:
    """Return the ffmpeg executable path to use in subprocess."""
    if not ffmpeg_location:
        return "ffmpeg"
    p = Path(ffmpeg_location)
    if p.is_file():
        return str(p)
    return str(p / "ffmpeg")


def trim_audio(input_path: Path, output_path: Path, intervals: list[tuple[float, float]], ffmpeg_location: str | None = None) -> bool:
    """
    Extract given intervals from input_path and write to output_path.
    If multiple intervals, concatenate them in order.
    """
    if not intervals:
        # no interval: copy whole file
        import shutil
        shutil.copy2(input_path, output_path)
        return True
    ffmpeg_bin = _ffmpeg_cmd(ffmpeg_location)
    if len(intervals) == 1:
        start, end = intervals[0]
        duration = end - start
        cmd = [
            ffmpeg_bin, "-y", "-i", str(input_path),
            "-ss", str(start), "-t", str(duration),
            "-c", "copy", str(output_path),
        ]
    else:
        # concat filter: [0]atrim=start1:end1[ a];[0]atrim=start2:end2[ b];[a][b]concat=n=2:v=0:a=1[out]
        filter_parts = []
        for i, (start, end) in enumerate(intervals):
            filter_parts.append(f"[0]atrim=start={start}:end={end},asetpts=PTS-STARTPTS[a{i}]")
        n = len(intervals)
        concat_inputs = "".join(f"[a{i}]" for i in range(n))
        filter_complex = ";".join(filter_parts) + ";" + concat_inputs + f"concat=n={n}:v=0:a=1[out]"
        cmd = [
            ffmpeg_bin, "-y", "-i", str(input_path),
            "-filter_complex", filter_complex, "-map", "[out]", str(output_path),
        ]
    try:
        subprocess.run(cmd, check=True, capture_output=True)
        return output_path.is_file()
    except subprocess.CalledProcessError as e:
        print(f"  ffmpeg error: {e}", file=sys.stderr)
        return False


def list_sheets(xlsx_path: Path) -> None:
    """Print sheet indices and names, then exit."""
    if not xlsx_path.is_file():
        sys.exit(f"Source file not found: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print("Sheets in", xlsx_path.name)
    for i, name in enumerate(wb.sheetnames):
        print(f"  -s {i}  {name!r}")
    wb.close()
    print("Use -s <ID> to select a sheet when running downloads.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Download Music Bingo audio from YouTube (sources/music_bingo_songs.xlsx). Use -s <ID> to select the sheet."
    )
    parser.add_argument(
        "-f", "--folder",
        metavar="FOLDER_NAME",
        help="Output folder name under output/ (e.g. -f my_event -> output/my_event/). Required unless using --list-sheets.",
    )
    parser.add_argument(
        "-s", "--sheet",
        type=int,
        default=0,
        metavar="ID",
        help="Sheet ID (0-based index) in the xlsx (default: 0). Use --list-sheets to see IDs.",
    )
    parser.add_argument(
        "--list-sheets",
        action="store_true",
        help="List sheet IDs and names from the xlsx, then exit (no download).",
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=SOURCES_XLSX,
        metavar="PATH",
        help="Path to music_bingo_songs.xlsx (default: sources/music_bingo_songs.xlsx)",
    )
    parser.add_argument(
        "-n", "--dry-run",
        action="store_true",
        help="Only print what would be downloaded, do not download",
    )
    parser.add_argument(
        "--ffmpeg-location",
        metavar="PATH",
        help="Path to ffmpeg binary or directory containing ffmpeg/ffprobe (if not in PATH).",
    )
    args = parser.parse_args()

    if args.list_sheets:
        list_sheets(args.source)
        return

    if not (args.folder and str(args.folder).strip()):
        parser.error("-f / --folder FOLDER_NAME is required for downloads (use --list-sheets to list sheets only)")

    ffmpeg_loc = args.ffmpeg_location.strip() if args.ffmpeg_location else None
    if not args.dry_run and not ffmpeg_loc:
        ffmpeg_bin = "ffmpeg"
        try:
            subprocess.run([ffmpeg_bin, "-version"], capture_output=True, check=True)
        except (subprocess.CalledProcessError, FileNotFoundError):
            print("ffmpeg not found. Install it (e.g. sudo apt install ffmpeg) or set --ffmpeg-location.", file=sys.stderr)
            sys.exit(1)

    out_dir = OUTPUT_BASE / args.folder.strip()
    if not args.dry_run:
        out_dir.mkdir(parents=True, exist_ok=True)

    try:
        rows = load_rows(args.source, args.sheet)
    except FileNotFoundError as e:
        sys.exit(str(e))
    except ValueError as e:
        sys.exit(str(e))

    if not rows:
        sys.exit("No rows with YouTube URLs found in the source file.")

    downloaded_count = 0
    for name, interval_str, url in rows:
        intervals = parse_interval(interval_str)
        if not intervals and interval_str:
            print(f"Skipping (invalid interval '{interval_str}'): {name}", file=sys.stderr)
            continue
        safe_name = sanitize_filename(name)
        out_path = out_dir / f"{safe_name}.m4a"
        if out_path.is_file() and not args.dry_run:
            print(f"Exists: {out_path.name}")
            continue
        if args.dry_run:
            print(f"Would download: {name} -> {out_path.name}  interval={interval_str or 'full'}")
            continue
        print(f"Downloading: {name}")
        tmp_file = None
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp_dir = Path(tmpdir)
                for attempt in range(1, DOWNLOAD_RETRIES + 1):
                    if attempt > 1:
                        print(f"  Retry {attempt}/{DOWNLOAD_RETRIES} in {DOWNLOAD_RETRY_DELAY_SEC}s…")
                        time.sleep(DOWNLOAD_RETRY_DELAY_SEC)
                    sys.stdout.write("  Connecting…" if attempt == 1 else f"  Retrying…")
                    sys.stdout.flush()
                    tmp_file = download_audio(url, tmp_dir, ffmpeg_loc)
                    if tmp_file:
                        break
                if not tmp_file:
                    print(f"  Failed after {DOWNLOAD_RETRIES} attempts: {url}", file=sys.stderr)
                    time.sleep(DOWNLOAD_PAUSE_BETWEEN_SEC)
                    continue
                if intervals:
                    if not trim_audio(tmp_file, out_path, intervals, ffmpeg_loc):
                        print(f"  Failed to trim: {out_path.name}", file=sys.stderr)
                        continue
                else:
                    import shutil
                    shutil.copy2(tmp_file, out_path)
                downloaded_count += 1
                print(f"  Saved: {out_path}")
                time.sleep(DOWNLOAD_PAUSE_BETWEEN_SEC)
        except KeyboardInterrupt:
            sys.stdout.write("\r" + " " * 40 + "\r")
            sys.stdout.flush()
            print("  Interrupted.")
            sys.exit(130)

    print(f"Done. Output directory: {out_dir}")
    if not args.dry_run:
        print(f"Songs downloaded: {downloaded_count}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(130)
