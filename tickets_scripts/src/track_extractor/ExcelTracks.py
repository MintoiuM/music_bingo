import openpyxl
import os

class ExcelTracks:
    def __init__(self, excel_path, sheet_index=0):
        self.tracks = dict()
        self.round_name = None
        self.get_playlist_tracks(excel_path, sheet_index)

    def get_playlist_tracks(self, excel_path, sheet_index):
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        workbook = openpyxl.load_workbook(excel_path)
        sheet_names = workbook.sheetnames

        if sheet_index < 0 or sheet_index >= len(sheet_names):
            raise IndexError(f"Sheet index {sheet_index} out of range. Available sheets: {sheet_names}")

        worksheet = workbook[sheet_names[sheet_index]]
        self.round_name = worksheet.title

        # Expecting columns: A=ID (number), B=Song Name, C=Artist(s) optional
        # Excel often stores numbers as float (1.0, 2.0) so we accept that
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            if row[0] is None and row[1] is None:
                continue
            # Allow ID in column A as int or float from Excel
            try:
                raw_id = row[0]
                if raw_id is None or raw_id == "":
                    continue
                track_id = int(float(raw_id)) if isinstance(raw_id, (int, float)) else int(raw_id)
            except (ValueError, TypeError):
                continue
            track_name = row[1]
            if track_name is None or (isinstance(track_name, str) and not track_name.strip()):
                continue
            track_name = str(track_name).strip()
            artist_names = str(row[2]).strip() if len(row) > 2 and row[2] else "Unknown"
            self.tracks[track_id] = {"id": track_id, "name": track_name, "artists": artist_names}

        n = len(self.tracks)
        print(f"📋 Loaded {n} tracks from sheet '{self.round_name}' (index {sheet_index}).")
        if n == 0:
            print("   → Check that column A = numeric ID (1,2,3...), column B = song name. Row 1 can be header.")
